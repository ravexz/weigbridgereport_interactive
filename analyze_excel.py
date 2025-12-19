import openpyxl

def analyze_excel(file_path):
    wb = openpyxl.load_workbook(file_path, keep_vba=True)
    print(f"Sheets: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"Max Row: {sheet.max_row}, Max Column: {sheet.max_column}")
        # Print first 20 rows to see structure
        for i, row in enumerate(sheet.iter_rows(max_row=20, max_col=10, values_only=True)):
            if any(cell is not None for cell in row):
                print(f"Row {i+1}: {row}")


if __name__ == "__main__":
    analyze_excel("Report.xlsx.xlsm")
