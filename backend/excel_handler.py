import openpyxl
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
try:
    import win32com.client as win32
except ImportError:
    win32 = None

from .database import get_all_entries


TEMPLATE_PATH = "Report.xlsx.xlsm"
OUTPUT_DIR = "PDF Records"

def update_excel_report(entries, report_date):
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        
    wb = openpyxl.load_workbook(TEMPLATE_PATH, keep_vba=True)
    sheet = wb["Report"]
    
    # Update Date
    # Column F is index 6 (1-indexed)
    sheet.cell(row=4, column=6).value = report_date
    
    # Mapping for zones to row starting points (based on CSV analysis)
    # Zone 1 Norah: Row 6
    # Zone 2: Row 9
    # Zone 3 Dennis: Row 15
    # Zone 4 Westone: Row 21
    
    zone_rows = {
        "ZONE 1 NORAH": 6,
        "ZONE 2": 9,
        "ZONE 3 DENNIS": 15,
        "ZONE 4 WESTONE": 21
    }
    
    # Sort entries by zone to ensure they are grouped
    # And then sort by clerk or just maintain input order within zone
    sorted_entries = sorted(entries, key=lambda x: x.get("zone", "").upper())
    
    current_row_idx = {k: v for k, v in zone_rows.items()}
    zone_totals = {k: 0 for k in zone_rows}
    
    # Clear ranges first (optional)
    for z, r_start in zone_rows.items():
        for r in range(r_start, r_start + 4): # Assume max 4 rows per zone for current layout
             for c in [5, 7, 8, 9, 10, 11, 12, 16, 17]:
                 sheet.cell(row=r, column=c).value = None

    for entry in sorted_entries:
        zone = entry.get("zone", "").upper().strip()
        match = None
        for k in zone_rows:
            # Handle template zone names with trailing spaces or different casing
            clean_k = k.upper().strip()
            if clean_k in zone or zone in clean_k:
                match = k
                break
        
        if match:
            r = current_row_idx[match]
            # Headers: ZONE(D), CLERK(E), (F empty), VEHICLE(G), ROUTE(H), (I empty), (J empty), FLD WGT(K), FACT WGT(L), VAR(M), EXP.VAR(N), (%)VAR(O), Scorch(P), Quality(Q)
            sheet.cell(row=r, column=5).value = entry.get("clerk")
            sheet.cell(row=r, column=7).value = entry.get("vehicle")
            sheet.cell(row=r, column=8).value = entry.get("route")
            sheet.cell(row=r, column=9).value = entry.get("time_out")
            sheet.cell(row=r, column=10).value = entry.get("time_in")
            sheet.cell(row=r, column=11).value = entry.get("fld_wgt")
            sheet.cell(row=r, column=12).value = entry.get("fact_wgt")
            sheet.cell(row=r, column=16).value = entry.get("scorch_kg")
            sheet.cell(row=r, column=17).value = entry.get("quality_pct")
            
            current_row_idx[match] += 1
            zone_totals[match] += (entry.get("fact_wgt") or 0)

    # TEMPORARILY DISABLED Hiding to check if it's causing the blank issue
    # for z, (r_start, r_end) in zone_ranges.items():
    #     for r in range(current_row_idx.get(z, r_start), r_end + 1):
    #         sheet.row_dimensions[r].hidden = True
    #     for r in range(r_start, current_row_idx.get(z, r_start)):
    #         sheet.row_dimensions[r].hidden = False

    # Update Summary Section (Rows 30+) from ALL historical data
    # This ensures the chart stays up-to-date
    all_history = get_all_entries()
    # Group by date then zone
    date_sums = {}
    for h in all_history:
        d = h['date']
        if d not in date_sums:
            date_sums[d] = {k: 0 for k in zone_rows}
        
        h_zone = h['zone'].upper()
        for k in zone_rows:
            if k.upper() in h_zone or h_zone in k.upper():
                date_sums[d][k] += (h['fact_wgt'] or 0)
                break
    
    # Sort dates
    sorted_dates = sorted(date_sums.keys())
    # Take last 20 dates to fit in summary table
    plot_dates = sorted_dates[-25:]
    
    # Clear summary table area
    for r in range(30, 56):
        for c in range(3, 8):
            sheet.cell(row=r, column=c).value = None
            
    for i, d in enumerate(plot_dates):
        row = 30 + i
        sheet.cell(row=row, column=3).value = d
        sheet.cell(row=row, column=4).value = date_sums[d]["ZONE 1 NORAH"]
        sheet.cell(row=row, column=5).value = date_sums[d]["ZONE 2"]
        sheet.cell(row=row, column=6).value = date_sums[d]["ZONE 3 DENNIS"]
        sheet.cell(row=row, column=7).value = date_sums[d]["ZONE 4 WESTONE"]

    output_filename = f"Report_{report_date.replace('-', '')}.xlsm"
    output_path = os.path.abspath(os.path.join(OUTPUT_DIR, output_filename))
    wb.save(output_path)
    
    pdf_path = output_path.replace(".xlsm", ".pdf")
    if win32:
        try:
            convert_to_pdf(output_path, pdf_path)
            return pdf_path
        except Exception as e:
            print(f"PDF conversion failed: {e}")
            return output_path
    
    return output_path

def convert_to_pdf(excel_path, pdf_path):
    import time
    import win32com.client as win32
    import pythoncom
    
    # Initialize COM for this thread
    pythoncom.CoInitialize()
    
    # Use os.path.abspath to ensure Excel gets the right path
    abs_xl_path = os.path.abspath(excel_path)
    abs_pdf_path = os.path.abspath(pdf_path)
    
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.EnableEvents = False # Disable macros/events
    try:
        wb = excel.Workbooks.Open(abs_xl_path)
        sheet = wb.Worksheets("Report")
        
        # Ensure focus is on the right sheet
        sheet.Activate()
        
        # Unhide all rows/columns to prevent blank output
        try:
            sheet.Rows.Hidden = False
            sheet.Columns.Hidden = False
        except:
            pass
            
        # Force Calculation and Refresh Charts to capture new openpyxl data
        excel.Calculation = -4105 # xlCalculationAutomatic
        excel.Calculate()
        wb.RefreshAll()
        
        # Wait for any background refreshes
        time.sleep(3)
        
        # Configure Page Setup for Landscape
        try:
            sheet.PageSetup.Orientation = 2 # xlLandscape
            sheet.PageSetup.Zoom = False
            sheet.PageSetup.FitToPagesWide = 1
            sheet.PageSetup.FitToPagesTall = 1
            # EXPLICIT PRINT AREA to force capture
            sheet.PageSetup.PrintArea = "$A$1:$Q$50" 
        except Exception as e:
            print(f"Warning: PageSetup adjustments failed: {e}")
        
        # EXPORT ONLY THE 'Report' SHEET TO PDF
        # 0 = xlTypePDF
        sheet.ExportAsFixedFormat(0, abs_pdf_path)
        print(f"PDF successfully exported to {abs_pdf_path}")
        
        wb.Close(False)
    except Exception as e:
        print(f"Error during PDF conversion: {e}")
        raise e
    finally:
        excel.Quit()


if __name__ == "__main__":
    # Test update
    sample_entries = [
        {"zone": "ZONE 1 NORAH", "clerk": "Test Clerk", "vehicle": "KXX001", "route": "R1", "fld_wgt": 100, "fact_wgt": 105, "scorch_kg": 5, "quality_pct": 25}
    ]
    path = update_excel_report(sample_entries, "2023-12-19")
    print(f"Generated test report: {path}")

