import openpyxl
import os
import sqlite3
from backend.excel_handler import update_excel_report

def verify_xlsx_writing():
    # 1. Get real data
    conn = sqlite3.connect("reports.db")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM daily_entries LIMIT 5")
    entries = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    if not entries:
        print("No entries to test.")
        return

    report_date = entries[0]['date']
    print(f"Testing for date: {report_date}")
    
    # 2. Update Excel
    result_path = update_excel_report(entries, report_date)
    print(f"Function returned: {result_path}")
    
    # Derrive XLSM path
    xl_path = result_path.replace(".pdf", ".xlsm")
    print(f"Verifying XLSM: {xl_path}")
    
    # 3. Read back with openpyxl
    wb = openpyxl.load_workbook(xl_path, data_only=True)
    sheet = wb["Report"]
    
    # Check some cells
    f4_val = sheet.cell(row=4, column=6).value
    d6_val = sheet.cell(row=6, column=4).value # Zone 1 Header (Hardcoded in template likely)
    e6_val = sheet.cell(row=6, column=5).value # Clerk
    k6_val = sheet.cell(row=6, column=11).value # Fld Wgt
    
    print(f"\n--- OPENPYXL READBACK ---")
    print(f"Cell F4 (Date): {f4_val}")
    print(f"Cell D6 (Zone): {d6_val}")
    print(f"Cell E6 (Clerk): {e6_val}")
    print(f"Cell K6 (Fld Wgt): {k6_val}")
    
    if not e6_val and not k6_val:
        print("\nFAILURE: Row 6 is empty even after openpyxl save!")
    else:
        print("\nSUCCESS: Data found in Excel file.")

if __name__ == "__main__":
    verify_xlsx_writing()
